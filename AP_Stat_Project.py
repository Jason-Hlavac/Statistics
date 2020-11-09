#Importing Modules
from openpyxl import load_workbook
import numpy as np
#Initializing directory
workbook = load_workbook(filename="Directory.xlsx")
workbook.sheetnames
sheet = workbook.active

#Initializing lists
treatment1 = []
treatment2= []
completeList = []
#function for pulling an email from the list and checking to see if it is used
def pullEmail():
    num = np.random.randint(2, sheet.max_row)
    if(num not in completeList):
        completeList.append(num)
        return(sheet.cell(row =num, column=4).value)
    elif(num in completeList):
        pullEmail()
    else:
        print("An Error occured in the pullEmail() function. Passing")
        pass

#Asking the user how many they want in the sample size
iterationCount = input("How many people do you want in your sample size")
#Loop that calls the pullEmail function for each treatment every other number starting with treatment 1.
for i in range(0, int(iterationCount)):
    if(i % 2 == 0):
        treatment1.append(pullEmail())
    elif(i % 2 == 1):
        treatment2.append(pullEmail())
    else:
        print("An Error occured in the main loop. Skipping i =" + str(i))
        
#Printing Results
print("Treatment 1:\n")
print("="*100 + "\n")
print(*treatment1, sep = " ")
print("="*100 + "\n")
print("Treatment 2:\n ")
print("="*100 + "\n")
print(*treatment2, sep = " ")
print("="*100 + "\n")

#Printing additional results to double check results
print("Treatment 1 is " + str(len(treatment1))+ " entries long")
print("Treatment 2 is " + str(len(treatment2))+ " entries long")
